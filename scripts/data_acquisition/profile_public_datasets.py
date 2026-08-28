"""Create lightweight structural profiles without copying dataset contents."""

from __future__ import annotations

import json
import zipfile
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from iia_benchmark.data import audit_pronto_archive, load_pronto_merged_csv


ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data" / "public_datasets"


def main() -> int:
    profile: dict[str, object] = {}
    piade_raw = DATA / "piade" / "raw_data.csv"
    if piade_raw.exists():
        frame = pd.read_csv(piade_raw, usecols=["equipment_ID", "alarm", "type"])
        alarms = frame["alarm"].dropna().astype(str)
        profile["piade_raw"] = {
            "rows": len(frame),
            "equipment": int(frame["equipment_ID"].nunique()),
            "alarm_codes_including_sentinel": int(alarms.nunique()),
            "non_sentinel_alarm_rows": int((alarms != "A_000").sum()),
            "machine_states": sorted(frame["type"].dropna().astype(str).unique().tolist()),
        }
    piade_sequences = DATA / "piade" / "sequences_1h_data.csv"
    if piade_sequences.exists():
        frame = pd.read_csv(piade_sequences)
        profile["piade_sequences"] = {
            "rows": len(frame),
            "columns": len(frame.columns),
            "equipment": int(frame["equipment_ID"].nunique()),
        }
    skab_data = DATA / "skab" / "data"
    if skab_data.exists():
        files = sorted(skab_data.rglob("*.csv"))
        profile["skab"] = {"csv_experiments": len(files)}
    tep_data = DATA / "tep_classic" / "data"
    if tep_data.exists():
        files = sorted(tep_data.glob("d*.dat"))
        profile["tep_classic"] = {
            "run_files": len(files),
            "normal_files": sum(path.stem.startswith("d00") for path in files),
            "fault_files": sum(not path.stem.startswith("d00") for path in files),
            "features": 52,
        }
    pronto_archive = DATA / "pronto" / "PRONTO_benchmark_case_study.zip"
    pronto_aligned = (
        DATA
        / "pronto/extracted/PRONTO benchmark case study/Pre-processed data/"
        "Aligned and labelled alarm and process data"
    )
    if pronto_archive.exists() and pronto_aligned.exists():
        paths = sorted(pronto_aligned.glob("Testday*_merged.csv"))
        runs = [load_pronto_merged_csv(path) for path in paths]
        archive = audit_pronto_archive(pronto_archive)
        label_counts = Counter(label for run in runs for label in run.labels.tolist())
        profile["pronto"] = {
            "archive_bytes": pronto_archive.stat().st_size,
            "archive_members": archive["members"],
            "archive_files": archive["files"],
            "archive_uncompressed_bytes": archive["uncompressed_bytes"],
            "archive_safe_to_extract": archive["safe_to_extract"],
            "aligned_test_days": len(runs),
            "aligned_samples": sum(len(run.labels) for run in runs),
            "alarm_tags_union": sorted(
                {name for run in runs for name in run.alarm_names}
            ),
            "process_features": list(runs[0].process_names) if runs else [],
            "fault_label_counts": dict(sorted(label_counts.items())),
            "validation": "finite process values, binary alarm states, nonempty labels",
        }
    fcc_alarm_archive = DATA / "fcc_alarm" / "alarmseriesdata.zip"
    fcc_timeseries_archive = DATA / "fcc_alarm" / "timeseriesdata.zip"
    if fcc_alarm_archive.exists() and fcc_timeseries_archive.exists():
        with zipfile.ZipFile(fcc_alarm_archive) as alarm_archive:
            alarm_files = sorted(
                name
                for name in alarm_archive.namelist()
                if name.lower().endswith(".csv")
            )
            alarm_categories = Counter(name.split("/")[1] for name in alarm_files)
            with alarm_archive.open(alarm_files[0]) as stream:
                alarm_sample = pd.read_csv(stream)
            alarm_crc_valid = alarm_archive.testzip() is None
            alarm_archive_entries = len(alarm_archive.infolist())
        with zipfile.ZipFile(fcc_timeseries_archive) as timeseries_archive:
            timeseries_files = sorted(
                name
                for name in timeseries_archive.namelist()
                if name.lower().endswith(".csv")
            )
            series_by_kind = Counter(
                Path(name).stem.rsplit("_", 1)[-1] for name in timeseries_files
            )
            sample_columns: dict[str, int] = {}
            sample_rows: dict[str, int] = {}
            for kind in ("process", "valves", "disturbances"):
                sample_name = next(
                    name
                    for name in timeseries_files
                    if Path(name).stem.endswith(f"_{kind}")
                )
                with timeseries_archive.open(sample_name) as stream:
                    sample = pd.read_csv(stream)
                sample_columns[kind] = len(sample.columns)
                sample_rows[kind] = len(sample)
            timeseries_crc_valid = timeseries_archive.testzip() is None
            timeseries_archive_entries = len(timeseries_archive.infolist())
        profile["fcc_alarm"] = {
            "simulation_runs": len(alarm_files),
            "abnormal_situations": len(alarm_categories),
            "runs_per_situation": dict(sorted(alarm_categories.items())),
            "samples_per_alarm_run": len(alarm_sample),
            "alarm_variables": len(alarm_sample.columns),
            "timeseries_csv_files": len(timeseries_files),
            "timeseries_files_by_kind": dict(sorted(series_by_kind.items())),
            "timeseries_sample_rows": sample_rows,
            "timeseries_columns_including_time": sample_columns,
            "alarm_archive_entries": alarm_archive_entries,
            "timeseries_archive_entries": timeseries_archive_entries,
            "zip_crc_valid": {
                "alarmseriesdata": alarm_crc_valid,
                "timeseriesdata": timeseries_crc_valid,
            },
            "representation_boundary": (
                "high-fidelity simulated FCC runs with scenario-directory labels; "
                "grouped evaluation must split by run and reserve complete abnormal "
                "situations for open-set testing"
            ),
        }
    comopi = DATA / "comopi" / "industrial_dataset_alarm_10m_agg.csv"
    if comopi.exists():
        rows = 0
        machines: set[str] = set()
        alarm_totals: Counter[str] = Counter()
        target_positive_bins = Counter({"AL_53": 0, "AL_54": 0})
        for chunk in pd.read_csv(comopi, chunksize=100_000):
            rows += len(chunk)
            machines.update(chunk["_serial"].astype(str).unique().tolist())
            alarms = chunk.filter(regex=r"^AL_")
            alarm_totals.update(
                {name: int(value) for name, value in alarms.sum().items()}
            )
            for target in target_positive_bins:
                target_positive_bins[target] += int((chunk[target] > 0).sum())
        profile["comopi_alarm_counts"] = {
            "rows": rows,
            "machines": len(machines),
            "bin_minutes": 10,
            "alarm_types": len(alarm_totals),
            "total_alarm_occurrences": sum(alarm_totals.values()),
            "target_positive_bins": dict(target_positive_bins),
            "top_alarm_counts": dict(alarm_totals.most_common(10)),
            "representation_boundary": (
                "within-bin event order is unavailable; AL_53/AL_54 are rare "
                "fault-condition targets, not flood-episode labels"
            ),
        }
    enas = DATA / "enas" / "EnAS_20200923-20210202.csv"
    if enas.exists():
        frame = pd.read_csv(enas)
        profile["enas_event_log"] = {
            "rows": len(frame),
            "columns": len(frame.columns),
            "start": str(frame["Timestamp"].iloc[0]),
            "end": str(frame["Timestamp"].iloc[-1]),
            "product_variant_counts": {
                str(key): int(value)
                for key, value in frame["PV"].value_counts().sort_index().items()
            },
            "representation_boundary": (
                "digital sensor/actuator state changes and manual error states; "
                "not an expert alarm-flood corpus"
            ),
        }
    imaks = DATA / "imaks" / "iMAKS_dataset.zip"
    if imaks.exists():
        with zipfile.ZipFile(imaks) as archive:
            with archive.open("sensors/timeseries_annotated.csv") as stream:
                frame = pd.read_csv(stream, low_memory=False)
            profile["imaks_synthetic"] = {
                "archive_entries": len(archive.infolist()),
                "annotated_sensor_rows": len(frame),
                "stations": int(frame["station_id"].nunique()),
                "sensors": int(frame["sensor_id"].nunique()),
                "anomaly_label_counts": {
                    str(key): int(value)
                    for key, value in frame["anomaly_label"].value_counts().items()
                },
                "alarm_flag_counts": {
                    str(key): int(value)
                    for key, value in frame["alarm_flag"].value_counts().items()
                },
                "representation_boundary": (
                    "synthetic data with alarm and causal truth; smoke/structural "
                    "validation only"
                ),
            }
    smd = DATA / "smd10towfgr" / "SCADA__monitoring_dataset_2020.xlsx"
    if smd.exists():
        workbook = load_workbook(smd, read_only=True, data_only=True)
        log_rows = 0
        distinct_codes: set[str] = set()
        event_types: Counter[str] = Counter()
        per_turbine: dict[str, int] = {}
        for sheet in workbook.worksheets:
            if not sheet.title.endswith("_logs.csv"):
                continue
            iterator = sheet.iter_rows(values_only=True)
            header = [str(value) if value is not None else "" for value in next(iterator)]
            code_index = header.index("Code")
            type_index = header.index("Event type")
            count = 0
            for row in iterator:
                if row[code_index] is None:
                    continue
                count += 1
                distinct_codes.add(str(row[code_index]))
                event_types[str(row[type_index])] += 1
            log_rows += count
            per_turbine[sheet.title.removesuffix("_logs.csv")] = count
        workbook.close()
        profile["smd10towfgr"] = {
            "turbines": len(per_turbine),
            "scada_rows_per_turbine_including_header": 26_209,
            "scada_features_including_timestamp": 132,
            "log_rows": log_rows,
            "distinct_event_codes": len(distinct_codes),
            "event_type_counts": dict(event_types.most_common()),
            "log_rows_by_turbine": per_turbine,
            "representation_boundary": (
                "timestamped alarm/event logs are directly usable for sequence and "
                "density tasks; flood-class labels require derivation or expert review"
            ),
        }
    output = DATA / "profile.json"
    output.write_text(json.dumps(profile, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(profile, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
