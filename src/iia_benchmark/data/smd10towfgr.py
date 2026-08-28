from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from .schema import AlarmEvent


DEFAULT_EVENT_TYPES = ("Alarm log (A)", "Warning log (W)")


def _timestamp_seconds(value: object) -> float | None:
    if value is None:
        return None
    parsed = pd.Timestamp(value) if isinstance(value, datetime) else pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed.value / 1_000_000_000)


def _event_code(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_smd_alarm_events(
    path: str | Path,
    *,
    turbines: Iterable[str] | None = None,
    event_types: Iterable[str] = DEFAULT_EVENT_TYPES,
) -> dict[str, tuple[AlarmEvent, ...]]:
    """Load timestamped Alarm/Warning occurrences from SMD10TOWFGR log sheets.

    Each source row is represented as an activation occurrence. Reset/ack fields are
    intentionally not converted into clearances here because several sheets contain
    mixed Excel/text encodings and missing reset values; downstream occurrence-bin
    validation must not invent standing-state semantics.
    """

    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(source)
    selected = None if turbines is None else {str(value).upper() for value in turbines}
    allowed = {str(value) for value in event_types}
    if not allowed:
        raise ValueError("event_types must not be empty")
    workbook = load_workbook(source, read_only=True, data_only=True)
    result: dict[str, tuple[AlarmEvent, ...]] = {}
    try:
        for sheet in workbook.worksheets:
            if not sheet.title.endswith("_logs.csv"):
                continue
            turbine = sheet.title.removesuffix("_logs.csv").upper()
            if selected is not None and turbine not in selected:
                continue
            rows = sheet.iter_rows(values_only=True)
            header = [str(value) if value is not None else "" for value in next(rows)]
            required = {"Code", "Detected", "Event type"}
            if not required.issubset(header):
                raise ValueError(f"{sheet.title} is missing required columns: {required - set(header)}")
            code_index = header.index("Code")
            detected_index = header.index("Detected")
            type_index = header.index("Event type")
            events: list[AlarmEvent] = []
            for row in rows:
                event_type = str(row[type_index])
                if event_type not in allowed or row[code_index] is None:
                    continue
                timestamp = _timestamp_seconds(row[detected_index])
                if timestamp is None:
                    continue
                priority = 1 if event_type == "Alarm log (A)" else 2
                events.append(
                    AlarmEvent(
                        timestamp=timestamp,
                        tag=_event_code(row[code_index]),
                        priority=priority,
                    )
                )
            result[turbine] = tuple(
                sorted(events, key=lambda event: (event.timestamp, event.tag, event.priority))
            )
    finally:
        workbook.close()
    if selected is not None and selected - set(result):
        raise ValueError(f"unknown turbines: {', '.join(sorted(selected - set(result)))}")
    return dict(sorted(result.items()))
